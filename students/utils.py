import pandas as pd
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import Student, CourseResult, StudentResult, Level, Course, Stage
from django.utils.translation import gettext_lazy as _
from datetime import datetime
import xlrd
import csv

def export_students_to_excel(students, academic_year=None):
    """
    تصدير بيانات الطلاب إلى ملف إكسل
    """
    # إنشاء دفتر عمل جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "قائمة الطلاب"
    
    # تعريف الألوان والأنماط
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    
    # إضافة عنوان الملف
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    if academic_year:
        title_cell.value = f"قائمة طلاب رواق العلوم الشرعية والعربية - العام الدراسي {academic_year}"
    else:
        title_cell.value = "قائمة طلاب رواق العلوم الشرعية والعربية"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # إعداد رؤوس الأعمدة
    headers = [
        'كود الطالب', 'رقم الجلوس', 'الاسم', 'النوع', 'الرقم القومي', 'المحافظة',
        'رقم الهاتف', 'المرحلة', 'المستوى', 'المذهب', 'نوع الدراسة', 'حالة القيد'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # إضافة بيانات الطلاب
    row_num = 4
    for student in students:
        row = [
            student.code,
            student.current_seat_number,
            student.full_name,
            student.gender,
            student.national_id,
            student.governorate,
            student.phone_number,
            student.level.stage.name,
            student.level.name,
            student.madhhab,
            student.study_type,
            student.enrollment_status
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        row_num += 1
    
    # ضبط عرض الأعمدة
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # إنشاء استجابة HTTP مع ملف الإكسل
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    if academic_year:
        filename = f"students_list_{academic_year}.xlsx"
    else:
        filename = f"students_list_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # حفظ الملف إلى استجابة HTTP
    wb.save(response)
    
    return response

def export_results_to_excel(student, academic_year=None):
    """
    تصدير بيان درجات طالب معين إلى ملف إكسل
    """
    # إنشاء دفتر عمل جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "بيان درجات"
    
    # تعريف الألوان والأنماط
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    pass_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    
    # إعداد معلومات الطالب
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "رواق العلوم الشرعية والعربية بالأزهر الشريف"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = "بيان درجات"
    subtitle_cell.font = Font(size=14, bold=True)
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # معلومات الطالب
    student_info = [
        ['اسم الطالب:', student.full_name, '', 'كود الطالب:', student.code],
        ['المستوى:', student.level.name, '', 'المرحلة:', student.level.stage.name],
        ['المذهب:', student.madhhab, '', 'نوع الدراسة:', student.study_type],
        ['رقم الجلوس:', student.current_seat_number, '', 'حالة القيد:', student.enrollment_status]
    ]
    
    for row_num, row_data in enumerate(student_info, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if col_num % 2 == 1:  # العناوين
                cell.font = Font(bold=True)
    
    # فلترة نتائج الطالب حسب العام الدراسي إذا تم تحديده
    if academic_year:
        course_results = CourseResult.objects.filter(student=student, academic_year=academic_year).order_by('course__name')
    else:
        course_results = CourseResult.objects.filter(student=student).order_by('-academic_year', 'course__name')
    
    # إضافة جدول النتائج
    headers = ['م', 'المادة', 'العام الدراسي', 'الدرجة', 'النهاية العظمى', 'الحالة']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row_num = 10
    for i, result in enumerate(course_results, 1):
        is_pass = result.score >= result.course.pass_score
        status = 'ناجح' if is_pass else 'راسب'
        
        row = [
            i,
            result.course.name,
            result.academic_year,
            result.score,
            result.course.max_score,
            status
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
            # تلوين خلية الحالة حسب النتيجة
            if col_num == 6:  # عمود الحالة
                cell.fill = pass_fill if is_pass else fail_fill
        
        row_num += 1
    
    # إضافة النتيجة النهائية
    final_results = StudentResult.objects.filter(student=student)
    if final_results.exists() and academic_year:
        final_result = final_results.filter(academic_year=academic_year).first()
        if final_result:
            ws.merge_cells(f'A{row_num+2}:F{row_num+2}')
            result_cell = ws[f'A{row_num+2}']
            result_cell.value = f"النتيجة النهائية: {final_result.result}"
            result_cell.font = Font(bold=True, size=12)
            result_cell.alignment = Alignment(horizontal='center')
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    
    # إنشاء استجابة HTTP مع ملف الإكسل
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"student_results_{student.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # حفظ الملف إلى استجابة HTTP
    wb.save(response)
    
    return response

def export_student_certificate(student):
    """
    تصدير إفادة قيد للطالب إلى ملف إكسل
    """
    # إنشاء دفتر عمل جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "إفادة قيد"
    
    # ضبط هوامش الصفحة
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # إعداد العنوان والشعار
    ws.merge_cells('A1:H1')
    header1 = ws['A1']
    header1.value = "الأزهر الشريف"
    header1.font = Font(size=16, bold=True)
    header1.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    header2 = ws['A2']
    header2.value = "رواق العلوم الشرعية والعربية"
    header2.font = Font(size=16, bold=True)
    header2.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:H4')
    title = ws['A4']
    title.value = "إفادة قيد طالب"
    title.font = Font(size=18, bold=True)
    title.alignment = Alignment(horizontal='center')
    
    # محتوى الإفادة
    ws.merge_cells('A6:H6')
    content1 = ws['A6']
    content1.value = "يشهد رواق العلوم الشرعية والعربية بالأزهر الشريف بأن:"
    content1.font = Font(size=12)
    content1.alignment = Alignment(horizontal='right')
    
    ws.merge_cells('A8:H8')
    name = ws['A8']
    name.value = f"الطالب/ {student.full_name}"
    name.font = Font(size=14, bold=True)
    name.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A10:H10')
    content2 = ws['A10']
    content2.value = f"مقيد بالرواق للعام الدراسي {datetime.now().year}/{datetime.now().year+1}م"
    content2.font = Font(size=12)
    content2.alignment = Alignment(horizontal='right')
    
    ws.merge_cells('A11:H11')
    content3 = ws['A11']
    content3.value = f"بالمستوى {student.level.name} - {student.level.stage.name}"
    content3.font = Font(size=12)
    content3.alignment = Alignment(horizontal='right')
    
    ws.merge_cells('A12:H12')
    content4 = ws['A12']
    content4.value = f"برقم جلوس: {student.current_seat_number}"
    content4.font = Font(size=12)
    content4.alignment = Alignment(horizontal='right')
    
    ws.merge_cells('A14:H14')
    content5 = ws['A14']
    content5.value = "وقد أعطيت له هذه الإفادة بناءً على طلبه لتقديمها إلى من يهمه الأمر."
    content5.font = Font(size=12)
    content5.alignment = Alignment(horizontal='right')
    
    # التاريخ
    ws.merge_cells('F16:H16')
    date_title = ws['F16']
    date_title.value = "تحريراً في:"
    date_title.font = Font(size=12, bold=True)
    date_title.alignment = Alignment(horizontal='right')
    
    ws.merge_cells('F17:H17')
    date_value = ws['F17']
    date_value.value = datetime.now().strftime("%Y/%m/%d")
    date_value.font = Font(size=12)
    date_value.alignment = Alignment(horizontal='center')
    
    # التوقيعات
    ws.merge_cells('F19:H19')
    signature1 = ws['F19']
    signature1.value = "مدير الرواق"
    signature1.font = Font(size=12, bold=True)
    signature1.alignment = Alignment(horizontal='center')
    
    # ضبط عرض الأعمدة
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 10
    
    # إنشاء استجابة HTTP مع ملف الإكسل
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"student_certificate_{student.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # حفظ الملف إلى استجابة HTTP
    wb.save(response)
    
    return response

def import_students_from_excel(file, update_existing=False):
    """
    استيراد بيانات الطلاب من ملف إكسل
    """
    try:
        df = pd.read_excel(file)
        
        # التحقق من وجود الأعمدة المطلوبة
        required_columns = ['كود الطالب', 'الاسم', 'النوع', 'رقم الجلوس', 'الرقم القومي', 'المرحلة', 'المستوى']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return False, f"الأعمدة التالية مفقودة: {', '.join(missing_columns)}"
        
        # بدء استيراد البيانات
        imported_count = 0
        updated_count = 0
        errors = []
        
        for _, row in df.iterrows():
            try:
                # البحث عن الطالب إذا كان موجوداً
                student = None
                student_code = str(row['كود الطالب']).strip()
                
                try:
                    student = Student.objects.get(code=student_code)
                    if not update_existing:
                        errors.append(f"الطالب {student_code} موجود بالفعل")
                        continue
                except Student.DoesNotExist:
                    student = Student()
                
                # العثور على المرحلة والمستوى
                stage_name = row['المرحلة']
                level_name = row['المستوى']
                
                try:
                    stage = Stage.objects.get(name=stage_name)
                    level = Level.objects.filter(stage=stage, name=level_name).first()
                    
                    if not level:
                        errors.append(f"المستوى {level_name} غير موجود في المرحلة {stage_name}")
                        continue
                    
                except Stage.DoesNotExist:
                    errors.append(f"المرحلة {stage_name} غير موجودة")
                    continue
                
                # بدء تعبئة بيانات الطالب
                student.code = student_code
                student.full_name = row['الاسم']
                student.gender = row['النوع']
                student.current_seat_number = str(row['رقم الجلوس'])
                student.national_id = str(row['الرقم القومي'])
                student.level = level
                
                # البيانات الاختيارية
                for col, field in [
                    ('المحافظة', 'governorate'),
                    ('رقم الهاتف', 'phone_number'),
                    ('المذهب', 'madhhab'),
                    ('نوع الدراسة', 'study_type'),
                    ('حالة القيد', 'enrollment_status'),
                    ('رقم الجلوس السابق', 'previous_seat_number')
                ]:
                    if col in df.columns and not pd.isna(row[col]):
                        setattr(student, field, row[col])
                
                # حفظ الطالب
                student.save()
                
                if student.id and student.id > 0:
                    if update_existing and student is not None:
                        updated_count += 1
                    else:
                        imported_count += 1
            
            except Exception as e:
                errors.append(f"خطأ في استيراد الطالب {row.get('كود الطالب', 'غير معروف')}: {str(e)}")
        
        # إعداد رسالة النجاح
        success_message = f"تم استيراد {imported_count} طالب جديد"
        if update_existing:
            success_message += f" وتحديث {updated_count} طالب"
        
        if errors:
            error_message = f"الأخطاء ({len(errors)}): {'; '.join(errors[:5])}"
            if len(errors) > 5:
                error_message += f" و{len(errors) - 5} أخطاء أخرى"
            return True, f"{success_message}. {error_message}"
        
        return True, success_message
    
    except Exception as e:
        return False, f"خطأ في استيراد الملف: {str(e)}"

def import_results_from_excel(file, academic_year):
    """
    استيراد نتائج الطلاب من ملف إكسل
    """
    try:
        df = pd.read_excel(file)
        
        # التحقق من وجود الأعمدة المطلوبة
        required_columns = ['كود الطالب', 'المادة', 'الدرجة']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return False, f"الأعمدة التالية مفقودة: {', '.join(missing_columns)}"
        
        # بدء استيراد النتائج
        imported_count = 0
        updated_count = 0
        errors = []
        
        for _, row in df.iterrows():
            try:
                # البحث عن الطالب
                student_code = str(row['كود الطالب']).strip()
                course_name = str(row['المادة']).strip()
                score = float(row['الدرجة'])
                
                try:
                    student = Student.objects.get(code=student_code)
                except Student.DoesNotExist:
                    errors.append(f"الطالب {student_code} غير موجود")
                    continue
                
                # البحث عن المادة
                try:
                    course = Course.objects.get(name=course_name, level=student.level)
                except Course.DoesNotExist:
                    errors.append(f"المادة {course_name} غير موجودة للمستوى {student.level.name}")
                    continue
                
                # البحث عن النتيجة إذا كانت موجودة مسبقًا
                course_result, created = CourseResult.objects.update_or_create(
                    student=student,
                    course=course,
                    academic_year=academic_year,
                    defaults={'score': score}
                )
                
                if created:
                    imported_count += 1
                else:
                    updated_count += 1
            
            except Exception as e:
                errors.append(f"خطأ في استيراد نتيجة للطالب {row.get('كود الطالب', 'غير معروف')}: {str(e)}")
        
        # إعداد رسالة النجاح
        success_message = f"تم استيراد {imported_count} نتيجة جديدة وتحديث {updated_count} نتيجة للعام الدراسي {academic_year}"
        
        if errors:
            error_message = f"الأخطاء ({len(errors)}): {'; '.join(errors[:5])}"
            if len(errors) > 5:
                error_message += f" و{len(errors) - 5} أخطاء أخرى"
            return True, f"{success_message}. {error_message}"
        
        return True, success_message
    
    except Exception as e:
        return False, f"خطأ في استيراد الملف: {str(e)}" 